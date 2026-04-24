import argparse
import math
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.properties import CalcProperties


WORKSHEET_NAME = "1D"
NETWORK_DIR = Path("MTGHP-core")
DEFAULT_OUTPUT_PATH = Path("steady_1d.xlsx")

INPUT_START_ROW = 5
SUMMARY_HEADER_ROW = 30
SUMMARY_START_ROW = 31
SOLVER_HEADER_ROW = 42
SOLVER_START_ROW = 43
SOLVER_STEPS = 0
TABLE_START_ROW = 42
WRITE_CHUNK_ROWS = 160
BOUNDARY_OPTIONS = ["Known WSE", "Known Depth", "Normal Depth", "Critical Depth", "Dead End", "Spill End", "Spill Zero"]

TOTAL_Q_REF = "$D$5"
N_REF = "$D$6"
SIDE_DEFAULT_REF = "$D$7"
KE_REF = "$D$8"
KC_REF = "$D$9"
KB_REF = "$D$10"
KJ_REF = "$D$11"
SPILL_COEFF_REF = "$D$12"
STAGE_RELAX_REF = "$D$13"
SPLIT_RELAX_REF = "$D$14"
JUNCTION_ENERGY_WT_REF = "$D$15"
MIN_DEPTH_REF = "$D$16"
DEFAULT_SLOPE_REF = "$D$17"
PROFILE_RELAX_REF = "$D$18"
PROFILE_TOL_REF = "$D$19"
RESET_REF = "$D$20"
INIT_J1_REF = "$D$21"
INIT_J2_REF = "$D$22"
INIT_A1_REF = "$D$23"
INIT_A2_REF = "$D$24"
SPILL_RELAX_REF = "$D$25"

G_REF = "$I$5"
RHO_REF = "$I$6"
FINAL_ALPHA1_REF = "$I$11"
FINAL_ALPHA2_REF = "$I$12"
FINAL_J1_REF = "$I$13"
FINAL_J2_REF = "$I$14"
TOTAL_SPILL_REF = "$I$15"

SUM_SEG = 1
SUM_FROM = 2
SUM_TO = 3
SUM_PTS = 4
SUM_LEN = 5
SUM_B0 = 6
SUM_B1 = 7
SUM_Z0 = 8
SUM_Z1 = 9
SUM_BED0 = 10
SUM_BED1 = 11
SUM_WS0 = 12
SUM_WS1 = 13
SUM_QIN = 14
SUM_SPILL = 15
SUM_QOUT = 16
SUM_YAVG = 17
SUM_SFAVG = 18
SUM_HF = 19
SUM_HJ = 20
SUM_HTOT = 21
SUM_MODE = 22
SUM_ANG0 = 23
SUM_ANG1 = 24

TABLE_SN = 1
TABLE_DIST = 2
TABLE_E = 3
TABLE_N = 4
TABLE_BED = 5
TABLE_WIDTH = 6
TABLE_SIDE = 7
TABLE_TYPE = 8
TABLE_DEF = 9
TABLE_DX = 10
TABLE_Q = 11
TABLE_YF = 12
TABLE_AF = 13
TABLE_RF = 14
TABLE_VF = 15
TABLE_WSF = 16
TABLE_EGF = 17
TABLE_SFF = 18
TABLE_FRF = 19
TABLE_MF = 20
TABLE_ERRF = 21
TABLE_DONEF = 22
TABLE_YB = 23
TABLE_AB = 24
TABLE_RB = 25
TABLE_VB = 26
TABLE_WSB = 27
TABLE_EGB = 28
TABLE_SFB = 29
TABLE_FRB = 30
TABLE_MB = 31
TABLE_ERRB = 32
TABLE_DONEB = 33
TABLE_Y = 34
TABLE_WS = 35
TABLE_EG = 36
TABLE_CREST_L = 37
TABLE_CREST_R = 38
TABLE_SPILL_L = 39
TABLE_SPILL_R = 40
TABLE_SPILL = 41
TABLE_CUM_SPILL = 42
TABLE_FR = 43
TABLE_REGIME = 44
TABLE_SF = 45


def col_letter(col_idx):
    result = ""
    while col_idx:
        col_idx, rem = divmod(col_idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def a1(row, col):
    return f"{col_letter(col)}{row}"


def set_cell(grid, row, col, value):
    grid[row - 1][col - 1] = value


def angle_deg(dx, dy):
    return math.degrees(math.atan2(dy, dx))


def angle_diff(a1_deg, a2_deg):
    diff = (a2_deg - a1_deg + 180.0) % 360.0 - 180.0
    return abs(diff)


def yes_no(value):
    if value is None:
        return False
    return str(value).strip().lower() in {"yes", "true", "1"}


def float_or_none(value):
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def canonical_bc_type(value):
    text = " ".join(str(value).strip().lower().split())
    if not text or text == "nan":
        return ""
    mapping = {
        "known ws": "Known WSE",
        "known wse": "Known WSE",
        "known depth": "Known Depth",
        "normal depth": "Normal Depth",
        "critical depth": "Critical Depth",
        "none": "Dead End",
        "dead end": "Dead End",
        "spill end": "Spill End",
        "spill zero": "Spill Zero",
    }
    return mapping.get(text, str(value).strip())


def fill_numeric(series, default):
    values = pd.to_numeric(series, errors="coerce")
    values = values.interpolate(limit_direction="both").bfill().ffill()
    return values.fillna(default)


def optional_series(df, column_name, default_value):
    if column_name in df.columns:
        return df[column_name]
    return pd.Series([default_value] * len(df), index=df.index)


def cumulative_distances(df):
    distances = [0.0]
    east = df["Easting"].astype(float).tolist()
    north = df["Northing"].astype(float).tolist()
    for idx in range(1, len(df)):
        dx = east[idx] - east[idx - 1]
        dy = north[idx] - north[idx - 1]
        distances.append(distances[-1] + math.hypot(dx, dy))
    return distances


def deflection_angles(df):
    east = df["Easting"].astype(float).tolist()
    north = df["Northing"].astype(float).tolist()
    angles = [0.0] * len(df)
    for idx in range(1, len(df) - 1):
        a0 = angle_deg(east[idx] - east[idx - 1], north[idx] - north[idx - 1])
        a1_deg = angle_deg(east[idx + 1] - east[idx], north[idx + 1] - north[idx])
        angles[idx] = angle_diff(a0, a1_deg)
    return angles


def clean_columns(df):
    cols = []
    seen = {}
    for idx, name in enumerate(df.columns):
        label = str(name).strip()
        if not label or label.lower().startswith("unnamed:"):
            label = f"blank_{idx}"
        if label in seen:
            seen[label] += 1
            label = f"{label}.{seen[label]}"
        else:
            seen[label] = 0
        cols.append(label)
    df = df.copy()
    df.columns = cols
    return df


def load_network():
    master = clean_columns(pd.read_excel(NETWORK_DIR / "master.xlsx", keep_default_na=False))
    from_col = "Upstream" if "Upstream" in master.columns else "From"
    to_col = "Downstream" if "Downstream" in master.columns else "To"
    up_bc_col = "Upstream BC" if "Upstream BC" in master.columns else None
    up_bc_value_col = "Upstream BC Value" if "Upstream BC Value" in master.columns else None
    dn_bc_col = "Downstream BC" if "Downstream BC" in master.columns else None
    dn_bc_value_col = "Downstream BC Value" if "Downstream BC Value" in master.columns else None
    segments = []

    for _, record in master.iterrows():
        name = str(record["Channel"]).strip()
        df = clean_columns(pd.read_csv(NETWORK_DIR / f"{name}.csv"))

        df["WidthFilled"] = fill_numeric(optional_series(df, "Bed width", 1.0), 1.0)
        df["SideFilled"] = fill_numeric(optional_series(df, "Bank slope", 0.0), 0.0)
        df["BedFilled"] = fill_numeric(optional_series(df, "Bed elevation", 0.0), 0.0)
        df["Chainage"] = cumulative_distances(df)
        df["Deflection"] = deflection_angles(df)

        left_to = optional_series(df, "Spill to", "")
        right_to = optional_series(df, "Spill to.1", "")
        left_crest = fill_numeric(optional_series(df, "Spillway Left Elevation", float("nan")), float("nan"))
        right_crest = fill_numeric(optional_series(df, "Spill Right Elevation", float("nan")), float("nan"))

        df["SpillLeftOn"] = optional_series(df, "Spill left", "").map(yes_no)
        df["SpillRightOn"] = optional_series(df, "Spill Right", "").map(yes_no)
        df["SpillLeftTo"] = left_to.fillna("").astype(str).str.strip()
        df["SpillRightTo"] = right_to.fillna("").astype(str).str.strip()
        df["SpillLeftCrest"] = left_crest
        df["SpillRightCrest"] = right_crest

        east = df["Easting"].astype(float).tolist()
        north = df["Northing"].astype(float).tolist()
        start_ang = angle_deg(east[1] - east[0], north[1] - north[0]) if len(df) > 1 else 0.0
        end_ang = angle_deg(east[-1] - east[-2], north[-1] - north[-2]) if len(df) > 1 else 0.0

        segments.append({
            "name": name,
            "from_node": str(record[from_col]).strip(),
            "to_node": str(record[to_col]).strip(),
            "df": df,
            "npts": len(df),
            "length": float(df["Chainage"].iloc[-1]) if len(df) else 0.0,
            "start_width": float(df["WidthFilled"].iloc[0]),
            "end_width": float(df["WidthFilled"].iloc[-1]),
            "start_side": float(df["SideFilled"].iloc[0]),
            "end_side": float(df["SideFilled"].iloc[-1]),
            "start_bed": float(df["BedFilled"].iloc[0]),
            "end_bed": float(df["BedFilled"].iloc[-1]),
            "start_angle": start_ang,
            "end_angle": end_ang,
            "upstream_bc_type": canonical_bc_type(record[up_bc_col]) if up_bc_col else "",
            "upstream_bc_value": float_or_none(record[up_bc_value_col]) if up_bc_value_col else None,
            "downstream_bc_type": canonical_bc_type(record[dn_bc_col]) if dn_bc_col else "",
            "downstream_bc_value": float_or_none(record[dn_bc_value_col]) if dn_bc_value_col else None,
            "junction_delta": 0.0,
        })

    junctions = {}
    for seg in segments:
        junctions.setdefault(seg["from_node"], {"incoming": [], "outgoing": []})["outgoing"].append(seg)
        junctions.setdefault(seg["to_node"], {"incoming": [], "outgoing": []})["incoming"].append(seg)

    for node, data in junctions.items():
        if node in ("Inlet", "Outlet", "Spill"):
            continue
        if not data["incoming"]:
            continue
        main = data["incoming"][0]
        for seg in data["outgoing"]:
            seg["junction_delta"] = angle_diff(main["end_angle"], seg["start_angle"])

    return segments


def boundary_label(seg_name, end_name):
    return f'{"Inlet" if end_name == "start" else "Outlet"}-{seg_name}'


def default_boundary_type(seg, end_name):
    if end_name == "start":
        return canonical_bc_type(seg.get("upstream_bc_type")) or "Normal Depth"
    return canonical_bc_type(seg.get("downstream_bc_type")) or "Normal Depth"


def default_boundary_value(seg, end_name):
    if end_name == "start":
        return seg.get("upstream_bc_value")
    return seg.get("downstream_bc_value")


def external_boundary_specs(segments, summary_rows):
    specs = []
    for seg in segments:
        if seg["from_node"].lower() == "inlet":
            specs.append({
                "label": boundary_label(seg["name"], "start"),
                "segment": seg["name"],
                "end": "start",
                "bc_type": default_boundary_type(seg, "start"),
                "bc_value": default_boundary_value(seg, "start"),
                "q_expr": a1(summary_rows[seg["name"]], SUM_QIN),
                "remarks": "Master inlet boundary" if seg.get("upstream_bc_type") else "Default inlet boundary from missing BC",
            })
        if seg["to_node"].lower() == "outlet":
            specs.append({
                "label": boundary_label(seg["name"], "end"),
                "segment": seg["name"],
                "end": "end",
                "bc_type": default_boundary_type(seg, "end"),
                "bc_value": default_boundary_value(seg, "end"),
                "q_expr": a1(summary_rows[seg["name"]], SUM_QOUT),
                "remarks": "Master outlet boundary" if seg.get("downstream_bc_type") else "Default outlet boundary from missing BC",
            })
    return specs


def spill_mapping_expr(segments_by_name, segment_table_rows, donor_names, target_name, target_idx):
    if target_idx < 2:
        return "0"
    seq = target_idx - 2
    terms = []
    for donor_name in donor_names:
        donor = segments_by_name[donor_name]
        donor_rows = []
        donor_df = donor["df"].reset_index(drop=True)
        for donor_idx, rec in donor_df.iterrows():
            left_match = yes_no(rec.get("SpillLeftOn", False)) and str(rec.get("SpillLeftTo", "")).strip().lower() == target_name.lower()
            right_match = yes_no(rec.get("SpillRightOn", False)) and str(rec.get("SpillRightTo", "")).strip().lower() == target_name.lower()
            if left_match or right_match:
                donor_rows.append(donor_idx)
        if seq < len(donor_rows):
            donor_row = segment_table_rows[donor_name]["data_start"] + donor_rows[seq]
            terms.append(a1(donor_row, TABLE_SPILL))
    return "0" if not terms else f"({' + '.join(terms)})"


def trap_area(width_ref, side_ref, depth_expr):
    return f"(({width_ref})+({side_ref})*({depth_expr}))*({depth_expr})"


def trap_perimeter(width_ref, side_ref, depth_expr):
    return f"({width_ref})+2*({depth_expr})*SQRT(1+({side_ref})^2)"


def trap_radius(width_ref, side_ref, depth_expr):
    area = trap_area(width_ref, side_ref, depth_expr)
    perimeter = trap_perimeter(width_ref, side_ref, depth_expr)
    return f"({area})/MAX({perimeter},1E-6)"


def trap_top_width(width_ref, side_ref, depth_expr):
    return f"({width_ref})+2*({side_ref})*({depth_expr})"


def velocity_expr(q_expr, area_expr):
    return f"ABS({q_expr})/MAX({area_expr},1E-6)"


def velocity_head_expr(v_expr):
    return f"(({v_expr})^2)/(2*{G_REF})"


def friction_slope_expr(q_expr, area_expr, radius_expr):
    return (
        f"((ABS({q_expr})*{N_REF})/"
        f"(MAX({area_expr},1E-6)*MAX(({radius_expr})^(2/3),1E-6)))^2"
    )


def froude_expr(q_expr, area_expr, top_width_expr):
    velocity = velocity_expr(q_expr, area_expr)
    return f"ABS({velocity})/SQRT({G_REF}*MAX(({area_expr})/MAX({top_width_expr},1E-6),1E-6))"


def specific_force_expr(width_ref, side_ref, depth_expr, q_expr, area_expr):
    return (
        f"(({width_ref})*({depth_expr})^2)/2"
        f"+({side_ref})*({depth_expr})^3/3"
        f"+(ABS({q_expr})^2)/({G_REF}*MAX({area_expr},1E-6))"
    )


def approx_critical_depth_expr(q_expr, width_ref):
    return f"MAX({MIN_DEPTH_REF},POWER((ABS({q_expr})^2)/({G_REF}*MAX({width_ref},0.1)^2),1/3))"


def branch_loss_expr(delta_deg, branch_q_expr, main_q_expr, vh_up_expr, vh_dn_expr):
    if delta_deg <= 0.0:
        return "0"
    sin_sq = math.sin(math.radians(delta_deg / 2.0)) ** 2
    return (
        f"({KJ_REF}*{sin_sq:.8f}"
        f"*(ABS({branch_q_expr})/MAX(ABS({main_q_expr}),1E-6))^2"
        f"*MAX({vh_up_expr},{vh_dn_expr}))"
    )


def summary_section_terms(summary_row, at_start, stage_expr, q_expr):
    width_ref = a1(summary_row, SUM_B0 if at_start else SUM_B1)
    side_ref = a1(summary_row, SUM_Z0 if at_start else SUM_Z1)
    bed_ref = a1(summary_row, SUM_BED0 if at_start else SUM_BED1)
    depth = f"MAX({MIN_DEPTH_REF},({stage_expr})-({bed_ref}))"
    area = trap_area(width_ref, side_ref, depth)
    radius = trap_radius(width_ref, side_ref, depth)
    velocity = velocity_expr(q_expr, area)
    vh = velocity_head_expr(velocity)
    sf = friction_slope_expr(q_expr, area, radius)
    force = specific_force_expr(width_ref, side_ref, depth, q_expr, area)
    return {
        "width": width_ref,
        "side": side_ref,
        "bed": bed_ref,
        "depth": depth,
        "area": area,
        "radius": radius,
        "velocity": velocity,
        "vh": vh,
        "sf": sf,
        "force": force,
    }


def summary_segment_terms(summary_row, q_expr, stage_up_expr, stage_dn_expr, junction_delta_deg=0.0, branch_q_expr="0", main_q_expr="1"):
    up = summary_section_terms(summary_row, True, stage_up_expr, q_expr)
    dn = summary_section_terms(summary_row, False, stage_dn_expr, q_expr)
    length_ref = a1(summary_row, SUM_LEN)
    a0_ref = a1(summary_row, SUM_ANG0)
    a1_ref = a1(summary_row, SUM_ANG1)
    hf = f"({length_ref})*(({up['sf']})+({dn['sf']}))/2"
    hve = (
        f"IF(({dn['velocity']})>({up['velocity']}),"
        f"{KC_REF}*(({dn['vh']})-({up['vh']})),"
        f"{KE_REF}*(({up['vh']})-({dn['vh']})))"
    )
    hbend = f"{KB_REF}*(ABS(({a1_ref})-({a0_ref}))/90)*MAX({up['vh']},{dn['vh']})"
    hj = branch_loss_expr(junction_delta_deg, branch_q_expr, main_q_expr, up["vh"], dn["vh"])
    hl = f"({hf})+({hve})+({hbend})+({hj})"
    return {"up": up, "dn": dn, "hf": hf, "hve": hve, "hbend": hbend, "hj": hj, "hl": hl}


def boundary_stage_formula(summary_row, end_label, type_ref, value_ref, q_ref):
    bed_ref = a1(summary_row, SUM_BED0 if end_label == "start" else SUM_BED1)
    width_ref = a1(summary_row, SUM_B0 if end_label == "start" else SUM_B1)
    z0_ref = a1(summary_row, SUM_BED0)
    z1_ref = a1(summary_row, SUM_BED1)
    length_ref = a1(summary_row, SUM_LEN)
    norm_depth = (
        f"MAX({MIN_DEPTH_REF},POWER("
        f"(ABS({q_ref})*{N_REF})/"
        f"(MAX({width_ref},0.1)*SQRT(MAX(ABS(({z1_ref}-{z0_ref})/MAX({length_ref},1E-6)),{DEFAULT_SLOPE_REF}))),"
        f"3/5))"
    )
    crit_depth = f"MAX({MIN_DEPTH_REF},POWER((ABS({q_ref})^2)/({G_REF}*MAX({width_ref},0.1)^2),1/3))"
    return (
        f'=IF(OR({type_ref}="Known WSE",{type_ref}="Known WS"),{value_ref},'
        f'IF({type_ref}="Known Depth",({bed_ref})+({value_ref}),'
        f'IF({type_ref}="Normal Depth",({bed_ref})+({norm_depth}),'
        f'IF({type_ref}="Critical Depth",({bed_ref})+({crit_depth}),'
        f'IF(OR({type_ref}="Dead End",{type_ref}="None",{type_ref}="Spill End"),IF({value_ref}="",({bed_ref})+{MIN_DEPTH_REF},{value_ref}),'
        f'IF({type_ref}="Spill Zero",({bed_ref})+{MIN_DEPTH_REF},'
        f'({bed_ref})+({norm_depth}))))))'
    )


def build_sheet():
    segments = load_network()
    segments_by_name = {seg["name"]: seg for seg in segments}
    segment_order = [seg["name"] for seg in segments]
    summary_rows = {name: SUMMARY_START_ROW + idx for idx, name in enumerate(segment_order)}
    boundary_specs = external_boundary_specs(segments, summary_rows)
    boundary_rows = {spec["label"]: 5 + idx for idx, spec in enumerate(boundary_specs)}

    segment_table_rows = {}
    current_row = TABLE_START_ROW
    for seg in segments:
        data_start = current_row + 3
        segment_table_rows[seg["name"]] = {
            "title": current_row,
            "header": current_row + 1,
            "units": current_row + 2,
            "data_start": data_start,
            "data_end": data_start + seg["npts"] - 1,
        }
        current_row = data_start + seg["npts"] + 2

    total_rows = current_row + 5
    total_cols = 45
    grid = [["" for _ in range(total_cols)] for _ in range(total_rows)]

    usproject_end_row = segment_table_rows["UsProject"]["data_end"]
    waterway_start_row = segment_table_rows["waterway"]["data_start"]
    waterway_end_row = segment_table_rows["waterway"]["data_end"]
    usexit_start_row = segment_table_rows["UsExit"]["data_start"]
    forebay_start_row = segment_table_rows["forebay"]["data_start"]
    side_start_row = segment_table_rows["sidechannel"]["data_start"]

    usexit_type_ref = a1(boundary_rows[boundary_label("UsExit", "end")], 14)
    forebay_type_ref = a1(boundary_rows[boundary_label("forebay", "end")], 14)
    side_type_ref = a1(boundary_rows[boundary_label("sidechannel", "end")], 14)
    usexit_dead_end_cond = f'OR({usexit_type_ref}="Dead End",{usexit_type_ref}="None")'
    forebay_dead_end_cond = f'OR({forebay_type_ref}="Dead End",{forebay_type_ref}="None")'
    side_dead_end_cond = f'OR({side_type_ref}="Dead End",{side_type_ref}="None")'

    current_alpha1 = f"IFERROR({FINAL_ALPHA1_REF},{INIT_A1_REF})"
    current_alpha2 = f"IFERROR({FINAL_ALPHA2_REF},{INIT_A2_REF})"
    current_j1 = f"IFERROR({FINAL_J1_REF},{INIT_J1_REF})"
    current_j2 = f"IFERROR({FINAL_J2_REF},{INIT_J2_REF})"

    j1_bed_ref = a1(summary_rows["UsProject"], SUM_BED1)
    j2_bed_ref = a1(summary_rows["waterway"], SUM_BED1)
    usexit_summary_q_ref = a1(summary_rows["UsExit"], SUM_QIN)
    waterway_summary_q_ref = a1(summary_rows["waterway"], SUM_QIN)
    forebay_summary_q_ref = a1(summary_rows["forebay"], SUM_QIN)
    side_summary_q_ref = a1(summary_rows["sidechannel"], SUM_QIN)

    us_end_ws_super_ref = a1(usproject_end_row, TABLE_WSF)
    us_end_m_super_ref = a1(usproject_end_row, TABLE_MF)
    ww_start_ws_sub_ref = a1(waterway_start_row, TABLE_WSB)
    ww_start_m_sub_ref = a1(waterway_start_row, TABLE_MB)
    ww_start_ws_final_ref = a1(waterway_start_row, TABLE_WS)
    ue_start_ws_sub_ref = a1(usexit_start_row, TABLE_WSB)
    ue_start_m_sub_ref = a1(usexit_start_row, TABLE_MB)
    ue_start_ws_final_ref = a1(usexit_start_row, TABLE_WS)
    ww_end_ws_super_ref = a1(waterway_end_row, TABLE_WSF)
    ww_end_m_super_ref = a1(waterway_end_row, TABLE_MF)
    fb_start_ws_sub_ref = a1(forebay_start_row, TABLE_WSB)
    fb_start_m_sub_ref = a1(forebay_start_row, TABLE_MB)
    fb_start_ws_final_ref = a1(forebay_start_row, TABLE_WS)
    sc_start_ws_sub_ref = a1(side_start_row, TABLE_WSB)
    sc_start_m_sub_ref = a1(side_start_row, TABLE_MB)
    sc_start_ws_final_ref = a1(side_start_row, TABLE_WS)
    usproject_done_ref = "$I$20"
    waterway_done_ref = "$I$21"
    usexit_done_ref = "$I$22"
    forebay_done_ref = "$I$23"
    side_done_ref = "$I$24"
    j1_ready_ref = "$I$25"
    j2_ready_ref = "$I$26"

    j1_sub_stage_actual = f"IFERROR({ww_start_ws_sub_ref},{INIT_J1_REF})"
    j1_sub_force_actual = f"IFERROR({ww_start_m_sub_ref},-1E99)"
    j1_super_stage_actual = f"IFERROR({us_end_ws_super_ref},{INIT_J1_REF})"
    j1_super_force_actual = f"IFERROR({us_end_m_super_ref},-1E99)"
    j1_target_actual = f"IF(({j1_super_force_actual})>=({j1_sub_force_actual}),({j1_super_stage_actual}),({j1_sub_stage_actual}))"
    j1_error_actual = f"(({j1_target_actual})-({current_j1}))"
    j1_branch_stage_delta = (
        f"(IFERROR({ue_start_ws_final_ref},IFERROR({ue_start_ws_sub_ref},{FINAL_J1_REF}))-"
        f"IFERROR({ww_start_ws_final_ref},IFERROR({ww_start_ws_sub_ref},{FINAL_J1_REF})))"
    )

    j2_sub_stage_actual = (
        f"IF(IFERROR({sc_start_m_sub_ref},-1E99)>=IFERROR({fb_start_m_sub_ref},-1E99),"
        f"IFERROR({sc_start_ws_sub_ref},IFERROR({fb_start_ws_sub_ref},{INIT_J2_REF})),"
        f"IFERROR({fb_start_ws_sub_ref},IFERROR({sc_start_ws_sub_ref},{INIT_J2_REF})))"
    )
    j2_sub_force_actual = f"MAX(IFERROR({fb_start_m_sub_ref},-1E99),IFERROR({sc_start_m_sub_ref},-1E99))"
    j2_super_stage_actual = f"IFERROR({ww_end_ws_super_ref},{INIT_J2_REF})"
    j2_super_force_actual = f"IFERROR({ww_end_m_super_ref},-1E99)"
    j2_target_actual = f"IF(({j2_super_force_actual})>=({j2_sub_force_actual}),({j2_super_stage_actual}),({j2_sub_stage_actual}))"
    j2_error_actual = f"(({j2_target_actual})-({current_j2}))"
    j2_branch_stage_delta = (
        f"(IFERROR({sc_start_ws_final_ref},IFERROR({sc_start_ws_sub_ref},{FINAL_J2_REF}))-"
        f"IFERROR({fb_start_ws_final_ref},IFERROR({fb_start_ws_sub_ref},{FINAL_J2_REF})))"
    )

    set_cell(grid, 1, 1, "1D Steady Open Channel Network Solver")
    set_cell(grid, 3, 1, "Two-pass steady 1D workbook: subcritical backwater, supercritical forewater, mixed-flow specific-force selection, and spill routing to drain.")

    set_cell(grid, 4, 2, "Parameter")
    set_cell(grid, 4, 3, "Unit")
    set_cell(grid, 4, 4, "Value")
    inputs = [
        ("Total inflow Q", "m3/s", 51),
        ("Manning n", "-", 0.015),
        ("Default bank slope", "H:V", 0.0),
        ("Expansion coeff Ke", "-", 0.30),
        ("Contraction coeff Kc", "-", 0.10),
        ("Bend coeff Kb", "-", 0.15),
        ("Junction coeff base", "-", 0.75),
        ("Spill coeff Cw", "-", 1.70),
        ("Stage relaxation", "-", 0.35),
        ("Split relaxation", "-", 0.08),
        ("Junction energy weight", "-", 0.05),
        ("Minimum depth", "m", 0.05),
        ("Boundary slope", "-", 0.001),
        ("Profile relaxation", "-", 0.15),
        ("Profile tolerance", "m", 0.001),
        ("Reset (0: Hold, 1: Eval)", "-", 1),
        ("J1 initial WSE", "m", 634.30),
        ("J2 initial WSE", "m", 631.90),
        ("alpha J1 initial", "-", 0.65),
        ("alpha J2 initial", "-", 0.65),
        ("Spill relaxation", "-", 0.20),
    ]
    for row_idx, (label, unit, value) in enumerate(inputs, start=INPUT_START_ROW):
        set_cell(grid, row_idx, 2, label)
        set_cell(grid, row_idx, 3, unit)
        set_cell(grid, row_idx, 4, value)

    set_cell(grid, 4, 8, "Fundamental Inputs")
    set_cell(grid, 5, 7, "Acceleration g")
    set_cell(grid, 5, 8, "m/s2")
    set_cell(grid, 5, 9, 9.81)
    set_cell(grid, 6, 7, "Density rho")
    set_cell(grid, 6, 8, "kg/m3")
    set_cell(grid, 6, 9, 998.2)
    set_cell(grid, 7, 7, "Allowed BC types")
    set_cell(grid, 7, 8, "-")
    set_cell(grid, 7, 9, "Known WSE / Known Depth / Normal Depth / Critical Depth / Dead End / Spill End / Spill Zero")
    set_cell(grid, 11, 7, "Final alpha J1")
    set_cell(grid, 11, 8, "-")
    set_cell(
        grid,
        11,
        9,
        f'=IF({RESET_REF}=0,{current_alpha1},IF({usexit_dead_end_cond},'
        f'MIN(0.98,MAX(0.02,MAX(0,1-({usexit_summary_q_ref}/MAX({TOTAL_Q_REF},1E-6))))),'
        f'MIN(0.98,MAX(0.02,{current_alpha1}+({SPLIT_RELAX_REF}*(({j1_branch_stage_delta})/MAX(ABS({FINAL_J1_REF}),1)))))))',
    )
    set_cell(grid, 12, 7, "Final alpha J2")
    set_cell(grid, 12, 8, "-")
    set_cell(
        grid,
        12,
        9,
        f'=IF({RESET_REF}=0,{current_alpha2},'
        f'IF(AND({forebay_dead_end_cond},{side_dead_end_cond}),'
        f'MIN(0.98,MAX(0.02,IFERROR({forebay_summary_q_ref}/MAX({waterway_summary_q_ref},1E-6),{current_alpha2}))),'
        f'IF({side_dead_end_cond},1,'
        f'IF({forebay_dead_end_cond},0,'
        f'MIN(0.98,MAX(0.02,{current_alpha2}+({SPLIT_RELAX_REF}*(({j2_branch_stage_delta})/MAX(ABS({FINAL_J2_REF}),1)))))))))',
    )
    set_cell(grid, 13, 7, "Final J1 WSE")
    set_cell(grid, 13, 8, "m")
    set_cell(
        grid,
        13,
        9,
        f'=IF({RESET_REF}=0,{current_j1},'
        f'IF(ABS($I$16)<{PROFILE_TOL_REF},({current_j1}),'
        f'MAX(({j1_bed_ref})+{MIN_DEPTH_REF},({current_j1})+({STAGE_RELAX_REF}*$I$16))))',
    )
    set_cell(grid, 14, 7, "Final J2 WSE")
    set_cell(grid, 14, 8, "m")
    set_cell(
        grid,
        14,
        9,
        f'=IF({RESET_REF}=0,{current_j2},'
        f'IF(ABS($I$17)<{PROFILE_TOL_REF},({current_j2}),'
        f'MAX(({j2_bed_ref})+{MIN_DEPTH_REF},({current_j2})+({STAGE_RELAX_REF}*$I$17))))',
    )
    set_cell(grid, 15, 7, "Total spill to drain")
    set_cell(grid, 15, 8, "m3/s")
    set_cell(grid, 15, 9, f"=IFERROR(N{summary_rows['drain']},0)")
    set_cell(grid, 16, 7, "J1 junction error")
    set_cell(grid, 16, 8, "m")
    set_cell(grid, 16, 9, f"=IFERROR($I$18-IFERROR($I$13,$D$21),0)")
    set_cell(grid, 17, 7, "J2 junction error")
    set_cell(grid, 17, 8, "m")
    set_cell(grid, 17, 9, f"=IFERROR($I$19-IFERROR($I$14,$D$22),0)")
    set_cell(grid, 18, 7, "J1 target WSE")
    set_cell(grid, 18, 8, "m")
    set_cell(grid, 18, 9, f"=IF({j1_ready_ref}=0,{current_j1},IFERROR({j1_target_actual},{INIT_J1_REF}))")
    set_cell(grid, 19, 7, "J2 target WSE")
    set_cell(grid, 19, 8, "m")
    set_cell(grid, 19, 9, f"=IF({j2_ready_ref}=0,{current_j2},IFERROR({j2_target_actual},{INIT_J2_REF}))")
    set_cell(grid, 20, 7, "UsProject done")
    set_cell(grid, 20, 8, "-")
    set_cell(grid, 20, 9, f"=IFERROR(MIN(V{segment_table_rows['UsProject']['data_start']}:V{segment_table_rows['UsProject']['data_end']},AG{segment_table_rows['UsProject']['data_start']}:AG{segment_table_rows['UsProject']['data_end']}),0)")
    set_cell(grid, 21, 7, "waterway done")
    set_cell(grid, 21, 8, "-")
    set_cell(grid, 21, 9, f"=IFERROR(MIN(V{segment_table_rows['waterway']['data_start']}:V{segment_table_rows['waterway']['data_end']},AG{segment_table_rows['waterway']['data_start']}:AG{segment_table_rows['waterway']['data_end']}),0)")
    set_cell(grid, 22, 7, "UsExit done")
    set_cell(grid, 22, 8, "-")
    set_cell(grid, 22, 9, f"=IFERROR(MIN(V{segment_table_rows['UsExit']['data_start']}:V{segment_table_rows['UsExit']['data_end']},AG{segment_table_rows['UsExit']['data_start']}:AG{segment_table_rows['UsExit']['data_end']}),0)")
    set_cell(grid, 23, 7, "forebay done")
    set_cell(grid, 23, 8, "-")
    set_cell(grid, 23, 9, f"=IFERROR(MIN(V{segment_table_rows['forebay']['data_start']}:V{segment_table_rows['forebay']['data_end']},AG{segment_table_rows['forebay']['data_start']}:AG{segment_table_rows['forebay']['data_end']}),0)")
    set_cell(grid, 24, 7, "sidechannel done")
    set_cell(grid, 24, 8, "-")
    set_cell(grid, 24, 9, f"=IFERROR(MIN(V{segment_table_rows['sidechannel']['data_start']}:V{segment_table_rows['sidechannel']['data_end']},AG{segment_table_rows['sidechannel']['data_start']}:AG{segment_table_rows['sidechannel']['data_end']}),0)")
    set_cell(grid, 25, 7, "J1 ready")
    set_cell(grid, 25, 8, "-")
    set_cell(grid, 25, 9, f"=IFERROR(MIN({usproject_done_ref},{waterway_done_ref},{usexit_done_ref}),0)")
    set_cell(grid, 26, 7, "J2 ready")
    set_cell(grid, 26, 8, "-")
    set_cell(grid, 26, 9, f"=IFERROR(MIN({waterway_done_ref},{forebay_done_ref},{side_done_ref}),0)")

    for col_idx, title in enumerate(["Boundary", "Segment", "End", "Type", "Value", "Resolved Stage", "Remarks"], start=11):
        set_cell(grid, 4, col_idx, title)

    for spec in boundary_specs:
        row_idx = boundary_rows[spec["label"]]
        bc_value = spec["bc_value"] if spec["bc_value"] is not None else ""
        set_cell(grid, row_idx, 11, spec["label"])
        set_cell(grid, row_idx, 12, spec["segment"])
        set_cell(grid, row_idx, 13, spec["end"])
        set_cell(grid, row_idx, 14, spec["bc_type"])
        set_cell(grid, row_idx, 15, bc_value)
        set_cell(
            grid,
            row_idx,
            16,
            boundary_stage_formula(summary_rows[spec["segment"]], spec["end"], a1(row_idx, 14), a1(row_idx, 15), spec["q_expr"]),
        )
        set_cell(grid, row_idx, 17, spec["remarks"])

    set_cell(grid, 26, 1, "Junction Internal Boundary Conditions")
    set_cell(grid, 27, 1, "Junction WSE cells iterate directly from the reach passes: forward-pass junction stage from the upstream reach and backward-pass junction stages from the connected downstream reaches are compared each recalculation, the junction error is relaxed in small increments, and the updated node stage is fed back into the reach passes.")

    summary_headers = [
        "Segment", "From", "To", "Pts", "Length",
        "b_start", "b_end", "z_start", "z_end",
        "bed_start", "bed_end", "WS_start", "WS_end",
        "Q_in", "Q_spill", "Q_out", "y_avg", "Sf_avg",
        "hf", "h_junc", "h_total", "Profile", "Start ang", "End ang",
    ]
    for col_idx, title in enumerate(summary_headers, start=1):
        set_cell(grid, SUMMARY_HEADER_ROW, col_idx, title)

    for seg in segments:
        row = summary_rows[seg["name"]]
        set_cell(grid, row, SUM_SEG, seg["name"])
        set_cell(grid, row, SUM_FROM, seg["from_node"])
        set_cell(grid, row, SUM_TO, seg["to_node"])
        set_cell(grid, row, SUM_PTS, seg["npts"])
        set_cell(grid, row, SUM_LEN, round(seg["length"], 3))
        set_cell(grid, row, SUM_B0, round(seg["start_width"], 3))
        set_cell(grid, row, SUM_B1, round(seg["end_width"], 3))
        set_cell(grid, row, SUM_Z0, round(seg["start_side"], 3))
        set_cell(grid, row, SUM_Z1, round(seg["end_side"], 3))
        set_cell(grid, row, SUM_BED0, round(seg["start_bed"], 4))
        set_cell(grid, row, SUM_BED1, round(seg["end_bed"], 4))
        set_cell(grid, row, SUM_MODE, "Mixed")
        set_cell(grid, row, SUM_ANG0, round(seg["start_angle"], 3))
        set_cell(grid, row, SUM_ANG1, round(seg["end_angle"], 3))

    usproject_row = summary_rows["UsProject"]
    waterway_row = summary_rows["waterway"]
    usexit_row = summary_rows["UsExit"]
    forebay_row = summary_rows["forebay"]
    side_row = summary_rows["sidechannel"]
    drain_row = summary_rows["drain"]
    usexit_seg = segments_by_name["UsExit"]
    forebay_seg = segments_by_name["forebay"]
    side_seg = segments_by_name["sidechannel"]
    usexit_type_ref = a1(boundary_rows[boundary_label("UsExit", "end")], 14)
    forebay_type_ref = a1(boundary_rows[boundary_label("forebay", "end")], 14)
    side_type_ref = a1(boundary_rows[boundary_label("sidechannel", "end")], 14)
    dead_end_pilot_q = "0.05"
    usexit_closed_q_expr = f"MAX({dead_end_pilot_q},IFERROR(O{usexit_row},0)+{dead_end_pilot_q})"
    forebay_closed_q_expr = f"MAX({dead_end_pilot_q},IFERROR(O{forebay_row},0)+{dead_end_pilot_q})"
    side_closed_q_expr = f"MAX({dead_end_pilot_q},IFERROR(O{side_row},0)+{dead_end_pilot_q})"

    qin_formulas = {
        "UsProject": f"={TOTAL_Q_REF}",
        "waterway": f'=IF({usexit_dead_end_cond},MAX(0,{TOTAL_Q_REF}-N{usexit_row}),{TOTAL_Q_REF}*{FINAL_ALPHA1_REF})',
        "UsExit": f'=IF({usexit_dead_end_cond},{usexit_closed_q_expr},MAX(0,{TOTAL_Q_REF}-N{waterway_row}))',
        "forebay": (
            f'=IF(AND({forebay_dead_end_cond},{side_dead_end_cond}),MAX(0,N{waterway_row}*{FINAL_ALPHA2_REF}),'
            f'IF({forebay_dead_end_cond},{forebay_closed_q_expr},'
            f'IF({side_dead_end_cond},MAX(0,N{waterway_row}-N{side_row}),N{waterway_row}*{FINAL_ALPHA2_REF})))'
        ),
        "sidechannel": (
            f'=IF(AND({forebay_dead_end_cond},{side_dead_end_cond}),MAX(0,N{waterway_row}-N{forebay_row}),'
            f'IF({side_dead_end_cond},{side_closed_q_expr},'
            f'IF({forebay_dead_end_cond},MAX(0,N{waterway_row}-N{forebay_row}),MAX(0,N{waterway_row}-N{forebay_row}))))'
        ),
        "drain": f"=O{forebay_row}+O{side_row}",
    }
    for seg_name, formula in qin_formulas.items():
        set_cell(grid, summary_rows[seg_name], SUM_QIN, formula)

    solver_headers = []
    for col_idx, title in enumerate(solver_headers, start=1):
        set_cell(grid, SOLVER_HEADER_ROW, col_idx, title)

    inlet_stage_ref = a1(boundary_rows[boundary_label("UsProject", "start")], 16)
    usexit_stage_ref = a1(boundary_rows[boundary_label("UsExit", "end")], 16)
    forebay_stage_ref = a1(boundary_rows[boundary_label("forebay", "end")], 16)
    side_stage_ref = a1(boundary_rows[boundary_label("sidechannel", "end")], 16)
    drain_start_stage_ref = a1(boundary_rows[boundary_label("drain", "start")], 16)
    drain_end_stage_ref = a1(boundary_rows[boundary_label("drain", "end")], 16)

    for idx in range(SOLVER_STEPS):
        row = SOLVER_START_ROW + idx
        prev_row = row - 1
        set_cell(grid, row, 1, idx)
        if row == SOLVER_START_ROW:
            set_cell(grid, row, 2, f"={INIT_A1_REF}")
            set_cell(grid, row, 3, f"={INIT_A2_REF}")
            set_cell(grid, row, 4, f"={INIT_J1_REF}")
            set_cell(grid, row, 5, f"={INIT_J2_REF}")
        else:
            set_cell(grid, row, 2, f"=Y{prev_row}")
            set_cell(grid, row, 3, f"=Z{prev_row}")
            set_cell(grid, row, 4, f"=AM{prev_row}")
            set_cell(grid, row, 5, f"=AN{prev_row}")

        set_cell(grid, row, 6, f"={TOTAL_Q_REF}")
        set_cell(grid, row, 7, f'=IF({usexit_type_ref}="None",MAX(0,F{row}-H{row}),MAX(0,F{row}*B{row}))')
        set_cell(grid, row, 8, f'=IF({usexit_type_ref}="None",{usexit_closed_q_expr},MAX(0,F{row}-G{row}))')
        set_cell(
            grid,
            row,
            9,
            f'=IF(AND({forebay_type_ref}="None",{side_type_ref}="None"),MAX(0,G{row}*C{row}),IF({forebay_type_ref}="None",{forebay_closed_q_expr},IF({side_type_ref}="None",MAX(0,G{row}-J{row}),MAX(0,G{row}*C{row}))))',
        )
        set_cell(
            grid,
            row,
            10,
            f'=IF(AND({forebay_type_ref}="None",{side_type_ref}="None"),MAX(0,G{row}-I{row}),IF({side_type_ref}="None",{side_closed_q_expr},IF({forebay_type_ref}="None",MAX(0,G{row}-I{row}),MAX(0,G{row}-I{row}))))',
        )
        set_cell(grid, row, 11, f"={inlet_stage_ref}")
        set_cell(grid, row, 12, f"={usexit_stage_ref}")
        set_cell(grid, row, 13, f"={forebay_stage_ref}")
        set_cell(grid, row, 14, f"={side_stage_ref}")
        usexit_solver_stage_ref = f'IF({usexit_type_ref}="None",AC{row},L{row})'
        forebay_solver_stage_ref = f'IF({forebay_type_ref}="None",AD{row},M{row})'
        side_solver_stage_ref = f'IF({side_type_ref}="None",AD{row},N{row})'

        up_super_terms = summary_segment_terms(usproject_row, f"F{row}", f"K{row}", f"AE{row}")
        ww_sub_terms = summary_segment_terms(
            waterway_row, f"G{row}", f"AC{row}", f"AD{row}",
            segments_by_name["waterway"]["junction_delta"], f"H{row}", f"F{row}"
        )
        ue_sub_terms = summary_segment_terms(
            usexit_row, f"H{row}", f"AC{row}", usexit_solver_stage_ref,
            segments_by_name["UsExit"]["junction_delta"], f"G{row}", f"F{row}"
        )
        ww_super_terms = summary_segment_terms(
            waterway_row, f"G{row}", f"AE{row}", f"AF{row}",
            segments_by_name["waterway"]["junction_delta"], f"H{row}", f"F{row}"
        )
        fb_sub_terms = summary_segment_terms(
            forebay_row, f"I{row}", f"AD{row}", forebay_solver_stage_ref,
            segments_by_name["forebay"]["junction_delta"], f"J{row}", f"G{row}"
        )
        sc_sub_terms = summary_segment_terms(
            side_row, f"J{row}", f"AD{row}", side_solver_stage_ref,
            segments_by_name["sidechannel"]["junction_delta"], f"I{row}", f"G{row}"
        )
        j1_sub_terms = summary_section_terms(usproject_row, False, f"AC{row}", f"F{row}")
        j1_super_terms = summary_section_terms(usproject_row, False, f"AE{row}", f"F{row}")
        j2_sub_terms = summary_section_terms(waterway_row, False, f"AD{row}", f"G{row}")
        j2_super_terms = summary_section_terms(waterway_row, False, f"AF{row}", f"G{row}")
        j1_sub_target = f'MAX(P{row},Q{row})'
        j2_sub_target = f'MAX(S{row},T{row})'
        j1_bed_ref = a1(usproject_row, SUM_BED1)
        j2_bed_ref = a1(waterway_row, SUM_BED1)

        set_cell(grid, row, 15, f"=K{row}+{up_super_terms['up']['vh']}-{up_super_terms['hl']}-{up_super_terms['dn']['vh']}")
        set_cell(grid, row, 16, f"=AD{row}+{ww_sub_terms['dn']['vh']}+{ww_sub_terms['hl']}-{ww_sub_terms['up']['vh']}")
        set_cell(grid, row, 17, f"=L{row}+{ue_sub_terms['dn']['vh']}+{ue_sub_terms['hl']}-{ue_sub_terms['up']['vh']}")
        set_cell(grid, row, 18, f"=AE{row}+{ww_super_terms['up']['vh']}-{ww_super_terms['hl']}-{ww_super_terms['dn']['vh']}")
        set_cell(grid, row, 19, f"=M{row}+{fb_sub_terms['dn']['vh']}+{fb_sub_terms['hl']}-{fb_sub_terms['up']['vh']}")
        set_cell(grid, row, 20, f"=N{row}+{sc_sub_terms['dn']['vh']}+{sc_sub_terms['hl']}-{sc_sub_terms['up']['vh']}")

        set_cell(
            grid,
            row,
            21,
            f"=IFERROR(((O{row})-({j1_sub_target}))/MAX(ABS(D{row}),1),0)",
        )
        set_cell(
            grid,
            row,
            22,
            f"=IFERROR(((R{row})-({j2_sub_target}))/MAX(ABS(E{row}),1),0)",
        )
        set_cell(
            grid,
            row,
            23,
            f"=ABS(({j1_sub_target})-D{row})",
        )
        set_cell(
            grid,
            row,
            24,
            f"=ABS(({j2_sub_target})-E{row})",
        )
        set_cell(
            grid,
            row,
            25,
            f'=IF({RESET_REF}=0,B{row},IFERROR(IF({usexit_type_ref}="None",1,MIN(0.98,MAX(0.02,B{row}+{SPLIT_RELAX_REF}*((Q{row}-P{row})/MAX(ABS(D{row}),1))))),B{row}))',
        )
        set_cell(
            grid,
            row,
            26,
            f'=IF({RESET_REF}=0,C{row},IFERROR(IF(AND({forebay_type_ref}="None",{side_type_ref}="None"),C{row},IF({side_type_ref}="None",1,IF({forebay_type_ref}="None",0,MIN(0.98,MAX(0.02,C{row}+{SPLIT_RELAX_REF}*((T{row}-S{row})/MAX(ABS(E{row}),1))))))),C{row}))',
        )
        set_cell(
            grid,
            row,
            29,
            f"=IF({RESET_REF}=0,D{row},IF(AC{row}=0,({j1_sub_target}),IF(ABS(({j1_sub_target})-AC{row})<{PROFILE_TOL_REF},AC{row},MAX(({j1_bed_ref})+{MIN_DEPTH_REF},AC{row}+{STAGE_RELAX_REF}*((({j1_sub_target})-AC{row}))))))",
        )
        set_cell(
            grid,
            row,
            30,
            f"=IF({RESET_REF}=0,E{row},IF(AD{row}=0,({j2_sub_target}),IF(ABS(({j2_sub_target})-AD{row})<{PROFILE_TOL_REF},AD{row},MAX(({j2_bed_ref})+{MIN_DEPTH_REF},AD{row}+{STAGE_RELAX_REF}*((({j2_sub_target})-AD{row}))))))",
        )
        set_cell(
            grid,
            row,
            31,
            f"=IF({RESET_REF}=0,D{row},"
            f"IF(AE{row}=0,O{row},"
            f"IF(ABS(U{row})<{PROFILE_TOL_REF},AE{row},"
            f"MAX(({j1_bed_ref})+{MIN_DEPTH_REF},"
            f"AE{row}+({STAGE_RELAX_REF}*(((O{row})-AE{row})+({JUNCTION_ENERGY_WT_REF}*U{row}*MAX(ABS(AE{row}-({j1_bed_ref})),1))))))))",
        )
        set_cell(
            grid,
            row,
            32,
            f"=IF({RESET_REF}=0,E{row},"
            f"IF(AF{row}=0,R{row},"
            f"IF(ABS(V{row})<{PROFILE_TOL_REF},AF{row},"
            f"MAX(({j2_bed_ref})+{MIN_DEPTH_REF},"
            f"AF{row}+({STAGE_RELAX_REF}*(((R{row})-AF{row})+({JUNCTION_ENERGY_WT_REF}*V{row}*MAX(ABS(AF{row}-({j2_bed_ref})),1))))))))",
        )
        set_cell(grid, row, 33, f"={j1_sub_terms['force']}")
        set_cell(grid, row, 34, f"={j1_super_terms['force']}")
        set_cell(grid, row, 35, f"={j2_sub_terms['force']}")
        set_cell(grid, row, 36, f"={j2_super_terms['force']}")
        set_cell(grid, row, 37, f'=IF(AH{row}>=AG{row},"Supercritical","Subcritical")')
        set_cell(grid, row, 38, f'=IF(AJ{row}>=AI{row},"Supercritical","Subcritical")')
        set_cell(grid, row, 27, f'=IF({RESET_REF}=0,D{row},IFERROR(D{row}+{STAGE_RELAX_REF}*(AC{row}-D{row}),D{row}))')
        set_cell(grid, row, 28, f'=IF({RESET_REF}=0,E{row},IFERROR(E{row}+{STAGE_RELAX_REF}*(AD{row}-E{row}),E{row}))')
        set_cell(grid, row, 39, f'=IF({RESET_REF}=0,D{row},IFERROR(D{row}+{STAGE_RELAX_REF}*(IF(AK{row}="Supercritical",AE{row},AC{row})-D{row}),D{row}))')
        set_cell(grid, row, 40, f'=IF({RESET_REF}=0,E{row},IFERROR(E{row}+{STAGE_RELAX_REF}*(IF(AL{row}="Supercritical",AF{row},AD{row})-E{row}),E{row}))')
        set_cell(grid, row, 41, f'=IF({RESET_REF}=0,0,IFERROR(ABS(IF(AK{row}="Supercritical",AE{row},AC{row})-D{row}),0))')
        set_cell(grid, row, 42, f'=IF({RESET_REF}=0,0,IFERROR(ABS(IF(AL{row}="Supercritical",AF{row},AD{row})-E{row}),0))')

    stage_inputs = {
        "UsProject": {"start": inlet_stage_ref, "end": FINAL_J1_REF},
        "waterway": {"start": FINAL_J1_REF, "end": FINAL_J2_REF},
        "UsExit": {"start": FINAL_J1_REF, "end": usexit_stage_ref},
        "forebay": {"start": FINAL_J2_REF, "end": forebay_stage_ref},
        "sidechannel": {"start": FINAL_J2_REF, "end": side_stage_ref},
        "drain": {"start": drain_start_stage_ref, "end": drain_end_stage_ref},
    }

    for seg in segments:
        if seg["to_node"].lower() != "outlet":
            continue
        end_bc_type = default_boundary_type(seg, "end")
        if end_bc_type not in {"Dead End", "None", "Spill End"}:
            continue
        boundary_row = boundary_rows[boundary_label(seg["name"], "end")]
        start_stage = stage_inputs[seg["name"]]["start"]
        seed_stage = f'IF({a1(boundary_row,15)}="",{start_stage},{a1(boundary_row,15)})'
        if seg["name"] == "UsExit":
            branch_start_ws_sub_ref = ue_start_ws_sub_ref
            junction_stage_ref = FINAL_J1_REF
            ready_ref = j1_ready_ref
        elif seg["name"] == "forebay":
            branch_start_ws_sub_ref = fb_start_ws_sub_ref
            junction_stage_ref = FINAL_J2_REF
            ready_ref = j2_ready_ref
        else:
            branch_start_ws_sub_ref = sc_start_ws_sub_ref
            junction_stage_ref = FINAL_J2_REF
            ready_ref = j2_ready_ref
        branch_mismatch = f"(({junction_stage_ref})-IFERROR({branch_start_ws_sub_ref},{junction_stage_ref}))"
        if end_bc_type in {"Dead End", "None"}:
            set_cell(
                grid,
                boundary_row,
                16,
                f"=IF({RESET_REF}=0,IFERROR(P{boundary_row},{seed_stage}),"
                f"IF({ready_ref}=0,IFERROR(P{boundary_row},{seed_stage}),"
                f"IF(({branch_mismatch})>{PROFILE_TOL_REF},"
                f"MAX(({seed_stage}),IFERROR(P{boundary_row},{seed_stage})+({STAGE_RELAX_REF}*({branch_mismatch}))),"
                f"IFERROR(P{boundary_row},{seed_stage}))))",
            )
            set_cell(grid, boundary_row, 17, "Dead End outlet; start from the master WSE and only raise the downstream stage until the branch backwater meets the solved junction stage.")
        else:
            bed_ref = a1(summary_rows[seg["name"]], SUM_BED1)
            set_cell(
                grid,
                boundary_row,
                16,
                f"=IF({RESET_REF}=0,IFERROR(P{boundary_row},{seed_stage}),"
                f"IF({ready_ref}=0,IFERROR(P{boundary_row},{seed_stage}),"
                f"IF(ABS({branch_mismatch})<{PROFILE_TOL_REF},IFERROR(P{boundary_row},{seed_stage}),"
                f"MAX(({bed_ref})+{MIN_DEPTH_REF},IFERROR(P{boundary_row},{seed_stage})+({STAGE_RELAX_REF}*({branch_mismatch}))))))",
            )
            set_cell(grid, boundary_row, 17, "Spill End outlet; start from the master WSE, relax spill from zero, then adjust the downstream stage up or down until the branch backwater matches the solved junction stage.")

    table_headers = [
        "SN", "Dist", "Easting", "Northing", "Bed Z", "Width", "Side z", "Type", "Def Ang", "dx",
        "Q_used", "y_super", "A_super", "R_super", "v_super", "WS_super", "EG_super", "Sf_super", "Fr_super", "M_super", "Err_super", "Done_super",
        "y_sub", "A_sub", "R_sub", "v_sub", "WS_sub", "EG_sub", "Sf_sub", "Fr_sub", "M_sub", "Err_sub", "Done_sub",
        "y_final", "WS_final", "EG_final", "Crest_L", "Crest_R", "Spill_L", "Spill_R", "Spill_Q", "Cum_spill",
        "Fr_final", "Regime", "Sf_final",
    ]
    table_units = [
        "-", "m", "m", "m", "m", "m", "-", "-", "deg", "m",
        "m3/s", "m", "m2", "m", "m/s", "m", "m", "-", "-", "m3", "m", "-",
        "m", "m2", "m", "m/s", "m", "m", "-", "-", "m3", "m", "-",
        "m", "m", "m", "m", "m", "m3/s", "m3/s", "m3/s", "m3/s", "-", "-", "-",
    ]

    for seg in segments:
        rows = segment_table_rows[seg["name"]]
        summary_row = summary_rows[seg["name"]]
        data_start = rows["data_start"]
        data_end = rows["data_end"]
        start_stage_expr = stage_inputs[seg["name"]]["start"]
        end_stage_expr = stage_inputs[seg["name"]]["end"]
        q_ref = a1(summary_row, SUM_QIN)
        mode_ref = a1(summary_row, SUM_MODE)
        end_bc_type = default_boundary_type(seg, "end")
        start_bc_type = default_boundary_type(seg, "start")
        is_spill_end_bc = end_bc_type == "Spill End"
        is_closed_bc = end_bc_type in {"Dead End", "None"} and seg["to_node"].lower() == "outlet"
        is_spill_zero_bc = start_bc_type == "Spill Zero" and seg["from_node"].lower() == "inlet"
        has_lateral_spill = bool(
            seg["df"]["SpillLeftOn"].any()
            or seg["df"]["SpillRightOn"].any()
        )
        uses_spill_pilot_flow = (
            (has_lateral_spill or is_spill_end_bc)
            and not is_closed_bc
        )
        q_floor_expr = dead_end_pilot_q if (is_closed_bc or is_spill_end_bc) else ("0.001" if uses_spill_pilot_flow else "0")

        set_cell(grid, rows["title"], 1, f"Segment: {seg['name']}")
        set_cell(grid, rows["title"], 5, f"=E{summary_row}")
        set_cell(grid, rows["title"], 6, "m length")
        set_cell(grid, rows["title"], 8, f"={q_ref}")
        set_cell(grid, rows["title"], 9, "m3/s")
        set_cell(grid, rows["title"], 11, f"={mode_ref}")
        set_cell(grid, rows["title"], 12, "profile")
        for col_idx, title in enumerate(table_headers, start=1):
            set_cell(grid, rows["header"], col_idx, title)
        for col_idx, unit in enumerate(table_units, start=1):
            set_cell(grid, rows["units"], col_idx, unit)

        df = seg["df"]
        for idx, (_, rec) in enumerate(df.iterrows()):
            row = data_start + idx
            prev_row = row - 1
            next_row = row + 1
            is_first = row == data_start
            is_last = row == data_end
            q_cell = a1(row, TABLE_Q)
            prev_q_cell = a1(prev_row, TABLE_Q)
            y_super_cell = a1(row, TABLE_YF)
            prev_y_super_cell = a1(prev_row, TABLE_YF)
            y_sub_cell = a1(row, TABLE_YB)
            next_y_sub_cell = a1(next_row, TABLE_YB)
            area_super_cell = a1(row, TABLE_AF)
            radius_super_cell = a1(row, TABLE_RF)
            vel_super_cell = a1(row, TABLE_VF)
            ws_super_cell = a1(row, TABLE_WSF)
            eg_super_cell = a1(row, TABLE_EGF)
            sf_super_cell = a1(row, TABLE_SFF)
            fr_super_cell = a1(row, TABLE_FRF)
            m_super_cell = a1(row, TABLE_MF)
            err_super_cell = a1(row, TABLE_ERRF)
            done_super_cell = a1(row, TABLE_DONEF)
            prev_done_super_cell = a1(prev_row, TABLE_DONEF)
            prev_fr_super_cell = a1(prev_row, TABLE_FRF)
            area_sub_cell = a1(row, TABLE_AB)
            radius_sub_cell = a1(row, TABLE_RB)
            vel_sub_cell = a1(row, TABLE_VB)
            ws_sub_cell = a1(row, TABLE_WSB)
            eg_sub_cell = a1(row, TABLE_EGB)
            sf_sub_cell = a1(row, TABLE_SFB)
            fr_sub_cell = a1(row, TABLE_FRB)
            m_sub_cell = a1(row, TABLE_MB)
            err_sub_cell = a1(row, TABLE_ERRB)
            done_sub_cell = a1(row, TABLE_DONEB)
            next_fr_sub_cell = a1(next_row, TABLE_FRB)
            next_done_sub_cell = a1(next_row, TABLE_DONEB)
            final_y_cell = a1(row, TABLE_Y)
            final_ws_cell = a1(row, TABLE_WS)
            final_eg_cell = a1(row, TABLE_EG)
            crest_l_cell = a1(row, TABLE_CREST_L)
            crest_r_cell = a1(row, TABLE_CREST_R)
            spill_l_cell = a1(row, TABLE_SPILL_L)
            spill_r_cell = a1(row, TABLE_SPILL_R)
            spill_cell = a1(row, TABLE_SPILL)
            cum_spill_cell = a1(row, TABLE_CUM_SPILL)
            prev_cum_spill_cell = a1(prev_row, TABLE_CUM_SPILL)
            final_fr_cell = a1(row, TABLE_FR)
            final_regime_cell = a1(row, TABLE_REGIME)
            final_sf_cell = a1(row, TABLE_SF)

            set_cell(grid, row, TABLE_SN, int(rec["SN"]))
            set_cell(grid, row, TABLE_DIST, round(float(rec["Chainage"]), 3))
            set_cell(grid, row, TABLE_E, round(float(rec["Easting"]), 3))
            set_cell(grid, row, TABLE_N, round(float(rec["Northing"]), 3))
            set_cell(grid, row, TABLE_BED, round(float(rec["BedFilled"]), 4))
            set_cell(grid, row, TABLE_WIDTH, round(float(rec["WidthFilled"]), 3))
            set_cell(grid, row, TABLE_SIDE, round(float(rec["SideFilled"]), 3))
            set_cell(grid, row, TABLE_TYPE, str(rec.get("Type", "")))
            set_cell(grid, row, TABLE_DEF, round(float(rec["Deflection"]), 3))
            set_cell(grid, row, TABLE_DX, 0 if is_first else f"={a1(row, TABLE_DIST)}-{a1(prev_row, TABLE_DIST)}")
            if is_spill_zero_bc and seg["name"] == "drain":
                if idx < 2:
                    set_cell(grid, row, TABLE_Q, 0)
                else:
                    mapped_expr = spill_mapping_expr(segments_by_name, segment_table_rows, [name for name in segments_by_name if name != "drain"], "drain", idx)
                    set_cell(grid, row, TABLE_Q, f"={prev_q_cell}+{mapped_expr}")
            elif is_first:
                set_cell(grid, row, TABLE_Q, f"={q_ref}")
            else:
                set_cell(grid, row, TABLE_Q, f"=MAX({q_floor_expr},{prev_q_cell}-{a1(prev_row, TABLE_SPILL)})")

            area_fwd = trap_area(a1(row, TABLE_WIDTH), a1(row, TABLE_SIDE), y_super_cell)
            radius_fwd = trap_radius(a1(row, TABLE_WIDTH), a1(row, TABLE_SIDE), y_super_cell)
            top_fwd = trap_top_width(a1(row, TABLE_WIDTH), a1(row, TABLE_SIDE), y_super_cell)
            vel_fwd = velocity_expr(q_cell, area_fwd)
            fr_fwd = froude_expr(q_cell, area_fwd, top_fwd)
            m_fwd = specific_force_expr(a1(row, TABLE_WIDTH), a1(row, TABLE_SIDE), y_super_cell, q_cell, area_fwd)
            ws_fwd = f"={a1(row, TABLE_BED)}+{y_super_cell}"
            eg_fwd = f"={ws_super_cell}+(({vel_fwd})^2)/(2*{G_REF})"
            sf_fwd = f"={friction_slope_expr(q_cell, area_fwd, radius_fwd)}"
            crit_depth = approx_critical_depth_expr(a1(row, TABLE_Q), a1(row, TABLE_WIDTH))
            safe_prev_y_super = f"IFERROR({prev_y_super_cell},{crit_depth})"
            safe_prev_done_super = f"IFERROR({prev_done_super_cell},0)"
            safe_prev_fr_super = f"IFERROR({prev_fr_super_cell},0)"
            safe_y_super = f"IFERROR({y_super_cell},0)"
            safe_fr_super = f"IFERROR({fr_super_cell},0)"
            safe_err_super = f"IFERROR({err_super_cell},0)"
            safe_next_y_sub = f"IFERROR({next_y_sub_cell},{crit_depth})"
            safe_next_done_sub = f"IFERROR({next_done_sub_cell},0)"
            safe_next_fr_sub = f"IFERROR({next_fr_sub_cell},0)"
            safe_y_sub = f"IFERROR({y_sub_cell},0)"
            safe_fr_sub = f"IFERROR({fr_sub_cell},0)"
            safe_err_sub = f"IFERROR({err_sub_cell},0)"
            safe_done_super = f"IFERROR({done_super_cell},0)"
            safe_done_sub = f"IFERROR({done_sub_cell},0)"
            safe_m_super = f"IFERROR({m_super_cell},-1E99)"
            safe_m_sub = f"IFERROR({m_sub_cell},-1E99)"
            super_to_crit_step = f"MIN(ABS({safe_err_super})*{PROFILE_RELAX_REF},ABS(({crit_depth})-{safe_y_super}))"
            super_to_crit_y = (
                f"IF(({crit_depth})>{safe_y_super},"
                f"MIN(({crit_depth}),{safe_y_super}+({super_to_crit_step})),"
                f"MAX(({crit_depth}),{safe_y_super}-({super_to_crit_step})))"
            )
            sub_to_crit_step = f"MIN(ABS({safe_err_sub})*{PROFILE_RELAX_REF},ABS(({crit_depth})-{safe_y_sub}))"
            sub_to_crit_y = (
                f"IF(({crit_depth})>{safe_y_sub},"
                f"MIN(({crit_depth}),{safe_y_sub}+({sub_to_crit_step})),"
                f"MAX(({crit_depth}),{safe_y_sub}-({sub_to_crit_step})))"
            )
            if is_first:
                set_cell(grid, row, TABLE_YF, f"=MAX({MIN_DEPTH_REF},({start_stage_expr})-E{row})")
                set_cell(grid, row, TABLE_ERRF, 0)
                set_cell(grid, row, TABLE_DONEF, 1)
            else:
                set_cell(
                    grid,
                    row,
                    TABLE_YF,
                    f"=IF({RESET_REF}=0,{safe_prev_y_super},"
                    f"IF({safe_prev_done_super}=0,{safe_prev_y_super},"
                    f"IF(OR({safe_y_super}=0,{safe_prev_fr_super}<1,{safe_fr_super}<1),{crit_depth},"
                    f"IF({safe_err_super}>0,"
                    f"MAX({MIN_DEPTH_REF},{safe_y_super}+(({safe_err_super})*{PROFILE_RELAX_REF}*IF(({safe_fr_super})<1,1,-1))),"
                    f"IF({safe_err_super}<0,"
                    f"IF(ABS(({crit_depth})-{safe_y_super})>{PROFILE_TOL_REF},"
                    f"MAX({MIN_DEPTH_REF},{super_to_crit_y}),"
                    f"{crit_depth}),"
                    f"MAX({MIN_DEPTH_REF},{safe_y_super}))))))",
                )
                local_fwd = (
                    f"IF({a1(prev_row, TABLE_VF)}>{vel_super_cell},{KE_REF}*(({a1(prev_row, TABLE_VF)}^2-{vel_super_cell}^2)/(2*{G_REF})),"
                    f"{KC_REF}*(({vel_super_cell}^2-{a1(prev_row, TABLE_VF)}^2)/(2*{G_REF})))"
                    f"+({KB_REF}*(ABS({a1(row, TABLE_DEF)})/90)*(MAX({a1(prev_row, TABLE_VF)}^2,{vel_super_cell}^2)/(2*{G_REF})))"
                )
                set_cell(
                    grid,
                    row,
                    TABLE_ERRF,
                    f"=IF({RESET_REF}=0,0,IFERROR({a1(prev_row, TABLE_EGF)}-((({a1(prev_row, TABLE_SFF)}+{sf_super_cell})/2)*{a1(row, TABLE_DX)})-({local_fwd})-{eg_super_cell},0))",
                )
                set_cell(
                    grid,
                    row,
                    TABLE_DONEF,
                    f'=IF({RESET_REF}=0,1,IFERROR(IF(ROUND({fr_super_cell},3)=1,1,IF(AND(ABS({err_super_cell})<{PROFILE_TOL_REF},{fr_super_cell}>1),1,0)),0))',
                )
            set_cell(grid, row, TABLE_AF, f"={area_fwd}")
            set_cell(grid, row, TABLE_RF, f"={radius_fwd}")
            set_cell(grid, row, TABLE_VF, f"={vel_fwd}")
            set_cell(grid, row, TABLE_WSF, ws_fwd)
            set_cell(grid, row, TABLE_EGF, eg_fwd)
            set_cell(grid, row, TABLE_SFF, sf_fwd)
            set_cell(grid, row, TABLE_FRF, f"={fr_fwd}")
            set_cell(grid, row, TABLE_MF, f"={m_fwd}")

            area_bwd = trap_area(a1(row, TABLE_WIDTH), a1(row, TABLE_SIDE), y_sub_cell)
            radius_bwd = trap_radius(a1(row, TABLE_WIDTH), a1(row, TABLE_SIDE), y_sub_cell)
            top_bwd = trap_top_width(a1(row, TABLE_WIDTH), a1(row, TABLE_SIDE), y_sub_cell)
            vel_bwd = velocity_expr(q_cell, area_bwd)
            fr_bwd = froude_expr(q_cell, area_bwd, top_bwd)
            m_bwd = specific_force_expr(a1(row, TABLE_WIDTH), a1(row, TABLE_SIDE), y_sub_cell, q_cell, area_bwd)
            ws_bwd = f"={a1(row, TABLE_BED)}+{y_sub_cell}"
            eg_bwd = f"={ws_sub_cell}+(({vel_bwd})^2)/(2*{G_REF})"
            sf_bwd = f"={friction_slope_expr(q_cell, area_bwd, radius_bwd)}"
            if is_last:
                set_cell(grid, row, TABLE_YB, f"=MAX({MIN_DEPTH_REF},({end_stage_expr})-E{row})")
                set_cell(grid, row, TABLE_ERRB, 0)
                set_cell(grid, row, TABLE_DONEB, 1)
            else:
                set_cell(
                    grid,
                    row,
                    TABLE_YB,
                    f"=IF({RESET_REF}=0,{safe_next_y_sub},"
                    f"IF({safe_next_done_sub}=0,{safe_next_y_sub},"
                    f"IF(OR({safe_y_sub}=0,{safe_next_fr_sub}>1),{crit_depth},"
                    f"IF({safe_err_sub}>0,"
                    f"MAX({MIN_DEPTH_REF},{safe_y_sub}+(({safe_err_sub})*{PROFILE_RELAX_REF}*IF(({safe_fr_sub})<1,1,-1))),"
                    f"IF({safe_err_sub}<0,"
                    f"IF(ABS(({crit_depth})-{safe_y_sub})>{PROFILE_TOL_REF},"
                    f"MAX({MIN_DEPTH_REF},{sub_to_crit_y}),"
                    f"{crit_depth}),"
                    f"MAX({MIN_DEPTH_REF},{safe_y_sub}))))))",
                )
                local_bwd = (
                    f"IF({vel_sub_cell}>{a1(next_row, TABLE_VB)},{KE_REF}*(({vel_sub_cell}^2-{a1(next_row, TABLE_VB)}^2)/(2*{G_REF})),"
                    f"{KC_REF}*(({a1(next_row, TABLE_VB)}^2-{vel_sub_cell}^2)/(2*{G_REF})))"
                    f"+({KB_REF}*(ABS({a1(next_row, TABLE_DEF)})/90)*(MAX({vel_sub_cell}^2,{a1(next_row, TABLE_VB)}^2)/(2*{G_REF})))"
                )
                set_cell(
                    grid,
                    row,
                    TABLE_ERRB,
                    f"=IF({RESET_REF}=0,0,IFERROR({a1(next_row, TABLE_EGB)}+((({sf_sub_cell}+{a1(next_row, TABLE_SFB)})/2)*{a1(next_row, TABLE_DX)})+({local_bwd})-{eg_sub_cell},0))",
                )
                set_cell(
                    grid,
                    row,
                    TABLE_DONEB,
                    f'=IF({RESET_REF}=0,1,IFERROR(IF(ROUND({fr_sub_cell},3)=1,1,IF(AND(ABS({err_sub_cell})<{PROFILE_TOL_REF},{fr_sub_cell}<1),1,0)),0))',
                )
            set_cell(grid, row, TABLE_AB, f"={area_bwd}")
            set_cell(grid, row, TABLE_RB, f"={radius_bwd}")
            set_cell(grid, row, TABLE_VB, f"={vel_bwd}")
            set_cell(grid, row, TABLE_WSB, ws_bwd)
            set_cell(grid, row, TABLE_EGB, eg_bwd)
            set_cell(grid, row, TABLE_SFB, sf_bwd)
            set_cell(grid, row, TABLE_FRB, f"={fr_bwd}")
            set_cell(grid, row, TABLE_MB, f"={m_bwd}")

            set_cell(
                grid,
                row,
                TABLE_Y,
                f'=IFERROR(IF({mode_ref}="Supercritical",MAX({MIN_DEPTH_REF},IFERROR({y_super_cell},{crit_depth})),IF({mode_ref}="Subcritical",MAX({MIN_DEPTH_REF},IFERROR({y_sub_cell},{crit_depth})),IF(AND({safe_done_super}=1,{safe_done_sub}=1),IF({safe_m_super}>={safe_m_sub},MAX({MIN_DEPTH_REF},IFERROR({y_super_cell},{crit_depth})),MAX({MIN_DEPTH_REF},IFERROR({y_sub_cell},{crit_depth}))),IF({safe_done_super}=1,MAX({MIN_DEPTH_REF},IFERROR({y_super_cell},{crit_depth})),IF({safe_done_sub}=1,MAX({MIN_DEPTH_REF},IFERROR({y_sub_cell},{crit_depth})),{crit_depth}))))),{crit_depth})',
            )
            final_area = trap_area(a1(row, TABLE_WIDTH), a1(row, TABLE_SIDE), final_y_cell)
            top_final = trap_top_width(a1(row, TABLE_WIDTH), a1(row, TABLE_SIDE), final_y_cell)
            vel_final = velocity_expr(q_cell, final_area)
            set_cell(grid, row, TABLE_WS, f"={a1(row, TABLE_BED)}+{final_y_cell}")
            set_cell(grid, row, TABLE_EG, f"={final_ws_cell}+(({vel_final})^2)/(2*{G_REF})")

            left_crest = ""
            if yes_no(rec.get("SpillLeftOn", False)) and str(rec.get("SpillLeftTo", "")).lower() == "drain" and not pd.isna(rec.get("SpillLeftCrest")):
                left_crest = round(float(rec["SpillLeftCrest"]), 3)
            right_crest = ""
            if yes_no(rec.get("SpillRightOn", False)) and str(rec.get("SpillRightTo", "")).lower() == "drain" and not pd.isna(rec.get("SpillRightCrest")):
                right_crest = round(float(rec["SpillRightCrest"]), 3)
            set_cell(grid, row, TABLE_CREST_L, left_crest)
            set_cell(grid, row, TABLE_CREST_R, right_crest)
            left_target_expr = f"{SPILL_COEFF_REF}*MAX({a1(row, TABLE_DX)},0.1)*POWER(MAX(0,{final_ws_cell}-{crest_l_cell}),1.5)"
            right_target_expr = f"{SPILL_COEFF_REF}*MAX({a1(row, TABLE_DX)},0.1)*POWER(MAX(0,{final_ws_cell}-{crest_r_cell}),1.5)"
            set_cell(
                grid,
                row,
                TABLE_SPILL_L,
                0 if left_crest == "" else (
                    f"=IF({RESET_REF}=0,IFERROR({spill_l_cell},0),"
                    f"IF(({left_target_expr})<=0,0,"
                    f"MAX(0,IFERROR({spill_l_cell},0)+({SPILL_RELAX_REF}*(({left_target_expr})-IFERROR({spill_l_cell},0))))))"
                ),
            )
            set_cell(
                grid,
                row,
                TABLE_SPILL_R,
                0 if right_crest == "" else (
                    f"=IF({RESET_REF}=0,IFERROR({spill_r_cell},0),"
                    f"IF(({right_target_expr})<=0,0,"
                    f"MAX(0,IFERROR({spill_r_cell},0)+({SPILL_RELAX_REF}*(({right_target_expr})-IFERROR({spill_r_cell},0))))))"
                ),
            )
            if is_spill_end_bc and is_last:
                set_cell(grid, row, TABLE_SPILL, f"={spill_l_cell}+{spill_r_cell}+MAX(0,{q_cell}-({spill_l_cell}+{spill_r_cell})-{q_floor_expr})")
            else:
                set_cell(grid, row, TABLE_SPILL, f"={spill_l_cell}+{spill_r_cell}")
            set_cell(grid, row, TABLE_CUM_SPILL, f"={spill_cell}" if is_first else f"={prev_cum_spill_cell}+{spill_cell}")
            set_cell(grid, row, TABLE_FR, f"={froude_expr(q_cell, final_area, top_final)}")
            set_cell(grid, row, TABLE_REGIME, f'=IF({final_fr_cell}<1,"Subcritical","Supercritical")')
            final_radius = trap_radius(a1(row, TABLE_WIDTH), a1(row, TABLE_SIDE), a1(row, TABLE_Y))
            set_cell(grid, row, TABLE_SF, f'={friction_slope_expr(q_cell, final_area, final_radius)}')

        set_cell(grid, summary_row, SUM_WS0, f"=IFERROR({a1(data_start, TABLE_WS)},IFERROR({start_stage_expr},0))")
        set_cell(grid, summary_row, SUM_WS1, f"=IFERROR({a1(data_end, TABLE_WS)},IFERROR({end_stage_expr},0))")
        set_cell(grid, summary_row, SUM_SPILL, "0" if seg["name"] == "drain" else f"=IFERROR(AGGREGATE(9,6,{a1(data_start, TABLE_SPILL)}:{a1(data_end, TABLE_SPILL)}),0)")
        if is_closed_bc or is_spill_end_bc:
            set_cell(grid, summary_row, SUM_QOUT, 0)
        else:
            set_cell(grid, summary_row, SUM_QOUT, f"=MAX(0,IFERROR({a1(data_end, TABLE_Q)},0)-IFERROR({a1(data_end, TABLE_SPILL)},0)-{q_floor_expr})")
        set_cell(grid, summary_row, SUM_YAVG, f"=IFERROR(AGGREGATE(1,6,{a1(data_start, TABLE_Y)}:{a1(data_end, TABLE_Y)}),0)")
        set_cell(grid, summary_row, SUM_SFAVG, f"=IFERROR(AGGREGATE(1,6,{a1(data_start, TABLE_SF)}:{a1(data_end, TABLE_SF)}),0)")
        set_cell(grid, summary_row, SUM_HF, f"=IFERROR(E{summary_row}*R{summary_row},0)")

    for seg in segments:
        row = summary_rows[seg["name"]]
        if seg["name"] == "waterway":
            terms = summary_segment_terms(row, a1(row, SUM_QIN), FINAL_J1_REF, FINAL_J2_REF, seg["junction_delta"], a1(usexit_row, SUM_QIN), TOTAL_Q_REF)
            set_cell(grid, row, SUM_HJ, f"={terms['hj']}")
        elif seg["name"] == "UsExit":
            terms = summary_segment_terms(row, a1(row, SUM_QIN), FINAL_J1_REF, f'IF({usexit_dead_end_cond},{usexit_stage_ref},{usexit_stage_ref})', seg["junction_delta"], a1(waterway_row, SUM_QIN), TOTAL_Q_REF)
            set_cell(grid, row, SUM_HJ, f"={terms['hj']}")
        elif seg["name"] == "forebay":
            terms = summary_segment_terms(row, a1(row, SUM_QIN), FINAL_J2_REF, forebay_stage_ref, seg["junction_delta"], a1(side_row, SUM_QIN), a1(waterway_row, SUM_QIN))
            set_cell(grid, row, SUM_HJ, f"={terms['hj']}")
        elif seg["name"] == "sidechannel":
            terms = summary_segment_terms(row, a1(row, SUM_QIN), FINAL_J2_REF, side_stage_ref, seg["junction_delta"], a1(forebay_row, SUM_QIN), a1(waterway_row, SUM_QIN))
            set_cell(grid, row, SUM_HJ, f"={terms['hj']}")
        else:
            set_cell(grid, row, SUM_HJ, 0)
        set_cell(grid, row, SUM_HTOT, f"=S{row}+T{row}")

    return grid, total_rows, total_cols


def fill_rgb(red, green, blue):
    return PatternFill(
        fill_type="solid",
        fgColor=f"{int(round(red * 255)):02X}{int(round(green * 255)):02X}{int(round(blue * 255)):02X}",
    )


def apply_fill(ws, start_row, end_row, start_col, end_col, fill, bold=False, font_size=None):
    font = Font(bold=bold, size=font_size) if bold or font_size else None
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.fill = fill
            if font is not None:
                cell.font = font


def write_grid_to_worksheet(ws, grid):
    for row_idx, row_values in enumerate(grid, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            if value == "":
                continue
            ws.cell(row=row_idx, column=col_idx, value=value)


def apply_boundary_dropdowns_excel(ws):
    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(BOUNDARY_OPTIONS) + '"',
        allow_blank=True,
    )
    validation.prompt = "Select a boundary condition type."
    validation.promptTitle = "Boundary Condition"
    validation.error = "Choose one of the allowed boundary-condition options."
    validation.errorTitle = "Invalid Boundary Condition"
    ws.add_data_validation(validation)
    validation.add(f"N5:N20")


def apply_excel_layout(ws, total_cols):
    ws.freeze_panes = "A30"
    ws.sheet_view.showGridLines = True
    ws.auto_filter.ref = f"A{SUMMARY_HEADER_ROW}:{col_letter(24)}{SUMMARY_START_ROW + 5}"

    for col_idx in range(1, total_cols + 1):
        letter = get_column_letter(col_idx)
        if col_idx in (1, 2, 3, 4):
            ws.column_dimensions[letter].width = 14
        elif col_idx <= 10:
            ws.column_dimensions[letter].width = 12
        elif col_idx <= 17:
            ws.column_dimensions[letter].width = 18
        else:
            ws.column_dimensions[letter].width = 13

    title_fill = fill_rgb(0.85, 0.91, 0.97)
    panel_fill = fill_rgb(0.96, 0.96, 0.96)
    boundary_fill = fill_rgb(0.95, 0.98, 0.92)
    summary_fill = fill_rgb(0.96, 0.94, 0.88)
    solver_fill = fill_rgb(0.93, 0.96, 0.99)

    apply_fill(ws, 1, 1, 1, 9, title_fill, bold=True, font_size=14)
    apply_fill(ws, 4, 26, 2, 4, panel_fill)
    apply_fill(ws, 4, 20, 11, 17, boundary_fill, bold=True)
    apply_fill(ws, SUMMARY_HEADER_ROW, SUMMARY_HEADER_ROW, 1, 24, summary_fill, bold=True)
    if SOLVER_STEPS > 0:
        apply_fill(ws, SOLVER_HEADER_ROW, SOLVER_HEADER_ROW, 1, 42, solver_fill, bold=True)


def configure_excel_calculation(workbook):
    workbook.calculation = CalcProperties(
        calcMode="auto",
        fullCalcOnLoad=True,
        forceFullCalc=True,
        iterate=True,
        iterateCount=250,
        iterateDelta=0.001,
    )


def write_1d_excel(output_path=DEFAULT_OUTPUT_PATH, worksheet_name=WORKSHEET_NAME):
    grid, total_rows, total_cols = build_sheet()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheet_name

    write_grid_to_worksheet(worksheet, grid)
    apply_boundary_dropdowns_excel(worksheet)
    apply_excel_layout(worksheet, total_cols)
    configure_excel_calculation(workbook)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Wrote Excel workbook to {output_path.resolve()}", flush=True)


def parse_args():
    parser = argparse.ArgumentParser(description="Build the 1D steady network workbook in Excel format.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH, help="Output Excel workbook path.")
    return parser.parse_args()


def main():
    args = parse_args()
    write_1d_excel(output_path=args.output)


if __name__ == "__main__":
    main()
